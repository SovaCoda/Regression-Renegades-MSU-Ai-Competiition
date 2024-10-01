# import os.path

# from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
# from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from datetime import datetime

# # If modifying these scopes, delete the file token.json.
# SCOPES = ["https://www.googleapis.com/auth/drive.metadata.readonly"]
# Credentials.from_authorized_user_file("token.json", SCOPES)
import pandas as pd
import os
import google.auth
from google.oauth2 import service_account
from googleapiclient.discovery import build
from openpyxl import load_workbook
from openpyxl.styles import Color
from openpyxl.formatting.rule import ColorScaleRule

# Set up the Drive API client
SCOPES = ['https://www.googleapis.com/auth/drive.readonly']

# Create credentials and build the service
credentials = Credentials.from_authorized_user_file("token.json", SCOPES)
service = build('drive', 'v3', credentials=credentials)

def find_folder_id(folder_name):
    query = f"mimeType='application/vnd.google-apps.folder' and name='{folder_name}' and trashed=false"
    results = service.files().list(q=query, fields="files(id, name)").execute()
    items = results.get('files', [])
    
    if not items:
        print(f'No folder found with the name: {folder_name}')
        return None
    else:
        return items[0]['id']  # Assuming the first match is the desired folder

def count_images_in_folder(folder_id):
    query = f"'{folder_id}' in parents and trashed=false"
    results = service.files().list(q=query, pageSize=1000, fields="files(id, name, mimeType, owners)").execute()
    files = results.get('files', [])
    
    image_count = 0
    user_image_counts = {}
    
    for file in files:
        # Check if the file is an image
        if file['mimeType'].startswith('image/'):
            image_count += 1
            owner = file['owners'][0]['emailAddress']
            user_image_counts[owner] = user_image_counts.get(owner, 0) + 1
            
    return image_count, user_image_counts

def traverse_folders(folder_id, folder_name, dataUserCounts):
    total_images = 0
    user_image_counts = {}
    
    # Count images in the current folder
    images_count, current_user_counts = count_images_in_folder(folder_id)
    total_images += images_count
    dataUserCounts["Buildings"].append(folder_name)
    dataUserCounts["Total Images"].append(images_count)
    if 'kacy365@gmail.com' in current_user_counts :
        dataUserCounts["Kacy"].append(current_user_counts['kacy365@gmail.com'])
    else : 
        dataUserCounts["Kacy"].append(0)

    if 'joreesesorrell@gmail.com' in current_user_counts:
        dataUserCounts["Reese"].append(current_user_counts['joreesesorrell@gmail.com'])
    else:
        dataUserCounts["Reese"].append(0)

    if 'jobc3194@gmail.com' in current_user_counts:
        dataUserCounts["Conner"].append(current_user_counts['jobc3194@gmail.com'])
    else:
        dataUserCounts["Conner"].append(0)
    
    # Update user counts
    for user, count in current_user_counts.items():
        user_image_counts[user] = user_image_counts.get(user, 0) + count

    # Store folder data
    folder_data = {'folder_name': folder_name, 'total_images': images_count}
    folder_data.update(current_user_counts)  # Add user counts to the folder data

    # Query for subfolders
    query = f"'{folder_id}' in parents and mimeType='application/vnd.google-apps.folder' and trashed=false"
    results = service.files().list(q=query, fields="files(id, name)").execute()
    folders = results.get('files', [])
    
    # Recursively traverse subfolders
    for folder in folders:
        subfolder_data = traverse_folders(folder['id'], folder['name'], dataUserCounts)
        total_images += subfolder_data['total_images']
        
        # Update folder data with user counts
        for user, count in subfolder_data.items():
            if user != 'total_images':
                folder_data[user] = folder_data.get(user, 0) + count
            
    return {**folder_data, 'total_images': total_images}

def apply_color_scale(filename):
    wb = load_workbook(filename)
    ws = wb.active
    
    # Apply 3-color scale to the second column (Total Images)
    color_scale = ColorScaleRule(
        start_type='min',
        start_color='FF0000',  # Red
        mid_type='percentile',        # Adjust this value as needed
        mid_value=15,
        mid_color='FFFF00',     # Yellow
        end_type='max',
        end_color='00FF00'      # Green
    )
    ws.conditional_formatting.add('B2:B11'.format(ws.max_row), color_scale)
    
    wb.save(filename)
    wb.close()

def main():
    # creds = None

    # if os.path.exists('token.json'):
    #     creds = Credentials.from_authorized_user_file('token.json', SCOPES)
    # # If there are no (valid) credentials available, let the user log in.
    # if not creds or not creds.valid:
    #     if creds and creds.expired and creds.refresh_token:
    #         creds.refresh(Request())
    #     else:
    #         flow = InstalledAppFlow.from_client_secrets_file(
    #             'credentials.json', SCOPES)
    #         creds = flow.run_local_server(port=0)
    #     # Save the credentials for the next run
    #     with open('token.json', 'w') as token:
    #         token.write(creds.to_json())

    folder_name = 'Ai_Competition'
    folder_id = find_folder_id(folder_name)
    
    if folder_id:
        dataUserCounts = {"Buildings" : [], "Total Images" : [], "Kacy" : [], "Reese" : [], "Conner" : []}
        data = []
        folder_data = traverse_folders(folder_id, folder_name, dataUserCounts)
        data.append(folder_data)

        # Create a DataFrame
        df = pd.DataFrame(dataUserCounts)

        # Fill missing values with 0
        df = df.fillna(0)
        df = df.iloc[1:]

        # Prepare column names
        

        # Calculate totals
        total_row = df.sum(numeric_only=True).to_frame().T
        df = pd.concat([df, total_row], ignore_index=True)
        
        now = datetime.now()
        dt_string = now.strftime("%d_%m_%Y_%H_%M")
        # Save to Excel
        df.to_excel(f'DataReports/building_image_counts_{dt_string}.xlsx', index=False)
        apply_color_scale(f'DataReports/building_image_counts_{dt_string}.xlsx')

        print(f'Excel file "building_image_counts_{dt_string}.xlsx" has been created.')

if __name__ == '__main__':
    main()